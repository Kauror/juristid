r"""Reading the four sources, without trusting any of them.

Every function here treats its input as evidence rather than as data: a
filename is a claim, a ZIP entry name is a claim, and a producer's spreadsheet
is somebody else's reading. Nothing is executed, nothing is unpacked to a path
the archive chose, and the only thing taken as identity is a SHA-256.

Three findings from the real corpus shaped this module.

**ZIP entry names arrive in two encodings and one of them is already broken.**
676 of 767 entries set the UTF-8 flag; 91 carry UTF-8 bytes without it, which
``zipfile`` mis-decodes as cp437. Separately, some *stored* names are mojibake
that predates the archive — cp437 bytes read as cp1252 before zipping, which
turned ``ä`` into ``„`` and deleted ``ü`` outright. The decoding is recorded per
entry so a damaged name cannot pass for a deliberate one.

**Never join the producer's rows to the archive by filename.** The KodaDash
workbook records a ``file_sha256`` per row and it binds 759 of 759 rows to the
archive exactly. The same data matched by (encoding-tolerant) filename produced
three collisions and five wrong assignments. Bytes first, always.

**A ZIP entry name is not a path until it is canonicalised.** Every entry in
the real archive is stored as ``Opinions\<name>.pdf`` — the producer wrote a
Windows separator into the member names. ``zipfile`` rewrites the OS separator
to ``/`` when it reads a name, so the same container reads as ``Opinions/x.pdf``
on Windows and ``Opinions\x.pdf`` on Linux. The separator is therefore
canonicalised here, before the safety guard and before anything is used as an
identity, so one archive means one path on every host.

**The archive's naming convention is a signal, not a date.** Every file is
``YYYY-MM-DD - Saaja - Pealkiri.pdf``, and that date is the letter's own date:
the register's VÄLJA falls on the same day 326 times and the next day 227 times.
It is parsed, kept, and never used as a sent date (Stage-2H brief 12, 13, 19).
"""

from __future__ import annotations

import datetime
import hashlib
import posixpath
import re
import unicodedata
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

#: `YYYY-MM-DD - Recipient - Title.pdf`, the opinions archive's own convention.
ARCHIVE_NAME = re.compile(r"^(\d{4})-(\d{2})-(\d{2})\s+-\s+(.*?)\s+-\s+(.+)\.pdf$", re.DOTALL)

#: A Riigikogu proceeding number — `662 SE`, `180 SE`, `21 OE`. An exact,
#: citable reference token rather than a word that happens to be shared.
LAW_REFERENCE = re.compile(r"\b(\d{1,4})\s*(SE|OE|UA)\b", re.IGNORECASE)

#: A Windows drive prefix — ``C:/x``, ``C:x``. Refused wherever it appears in an
#: archive member name, on every host: it is not a relative path, and a reader
#: that treats it as one is one `os.path.join` away from writing outside the
#: directory it was given.
WINDOWS_DRIVE = re.compile(r"^[A-Za-z]:")

#: File signatures the archive is allowed to contain. The corpus is 767 PDFs;
#: anything else is reported rather than quietly catalogued as an opinion.
SIGNATURES: tuple[tuple[bytes, str], ...] = (
    (b"%PDF", "application/pdf"),
    (b"PK\x03\x04", "application/zip"),
    (b"\xd0\xcf\x11\xe0", "application/x-ole-storage"),
)

#: Words that appear in half the register and therefore distinguish nothing.
#:
#: Written as the *Estonian words themselves* and folded below, rather than as
#: hand-typed ASCII skeletons. Writing the skeletons by hand is what produced
#: the defect this replaces: the list carried ``poordumine`` and ``prdumine``
#: for *pöördumine*, and `fold` produces neither. It replaces each ``ö`` with a
#: space, so the real token is ``rdumine`` — seven characters, in no stopword,
#: and therefore accepted as a "distinctive" word. It appears in 207 register
#: titles, and it is the entire third signal under a link production holds
#: today between "Pöördumine seoses õigusloome ja bürokraatia vähendamisega"
#: and "Väljatöötamiskavatsustega seotud pöördumine" — two different subjects.
#:
#: Folding the list at import is what makes the intent and the effect the same
#: thing. A word can now be added in the spelling a reader recognises, and
#: whether `fold` mangles it stops being something the author has to predict.
_STOPWORD_SOURCE: tuple[str, ...] = (
    # Structural words: what the department *did*, never what about.
    "arvamus",
    "arvamuse",
    "eelnõu",
    "esitamine",
    "esitamise",
    "ettepanek",
    "ettepanekud",
    "kohta",
    "määrus",
    "määruse",
    "muudatused",
    "muutmine",
    "muutmise",
    "pöördumine",
    "pöördumised",
    "seadus",
    "seaduse",
    "seaduste",
    "seonduvalt",
    "sisend",
    "teiste",
    "vtk",
    # Process and institution words the register writes on hundreds of
    # unrelated rows. Each was measured against the real register before being
    # added, and each was found carrying an automatic link on its own:
    # "komisjoni" on 394 rows, "konsultatsioon" 396, "valitsuse" 204,
    # "euroopa" 746, "direktiivi" 436.
    "avalik",
    "direktiivi",
    "euroopa",
    "komisjoni",
    "konsultatsioon",
    "sellega",
    "tingimused",
    "vabariigi",
    "valitsuse",
)


#: How long a word has to be before sharing it says anything. Estonian
#: compounds are long, so a seven-character shared word is a real token
#: ("maksejuetuse", "autoriiguse") rather than a coincidence.
MINIMUM_TITLE_TOKEN = 7


class OpinionSourceError(RuntimeError):
    """A source is not what it claimed to be."""


def fold(value: object) -> str:
    """The comparison key used everywhere in this stage.

    Casefold, replace every non-ASCII character with a space, keep
    alphanumerics. Replacing rather than transliterating is deliberate: the
    archive destroyed some diacritics outright — cp1252 has no code point for
    cp437's ``ü``, so the letter is simply gone — and a transliteration would
    have to invent it back. Doing the same thing on both sides converges the
    damaged and undamaged spellings of one name without adding a character to
    either, at the cost of splitting a compound where a diacritic was; that
    cost is symmetric, which is what matters.
    """
    text = unicodedata.normalize("NFC", str(value or "")).casefold()
    text = "".join(character if character.isascii() else " " for character in text)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


#: Estonian letters, transliterated. Used *only* by `keyword_fold`.
TRANSLITERATION = str.maketrans({"ä": "a", "ö": "o", "õ": "o", "ü": "u", "š": "s", "ž": "z"})


def keyword_fold(value: object) -> str:
    """Like `fold`, but transliterating, for looking up words we chose ourselves.

    The distinction is the whole reason there are two functions. `fold` compares
    two *sources* and must not invent a character the archive destroyed, so it
    replaces rather than transliterates — which splits `ühispöördumine` into
    fragments. That is fine when both sides split identically and useless when
    one side is a word this codebase wrote down. Transliteration is safe here
    because nothing is being identified: a keyword either appears or it does
    not, and a wrong answer costs a submission kind rather than a Matter.
    """
    text = unicodedata.normalize("NFC", str(value or "")).casefold().translate(TRANSLITERATION)
    text = "".join(character if character.isascii() else " " for character in text)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _folded_stopwords(words: tuple[str, ...]) -> frozenset[str]:
    """Every token `fold` actually produces from the words above.

    A word that folds to nothing long enough to be a title token — ``määruse``
    becomes ``m ruse`` — contributes nothing and is simply absent, which is
    correct: `title_tokens` could never have emitted it either.
    """
    return frozenset(
        token for word in words for token in fold(word).split() if len(token) >= MINIMUM_TITLE_TOKEN
    )


TITLE_STOPWORDS: frozenset[str] = _folded_stopwords(_STOPWORD_SOURCE)


def title_tokens(value: object) -> frozenset[str]:
    """Distinctive words in a title, for the third exact signal."""
    return frozenset(
        word
        for word in fold(value).split()
        if len(word) >= MINIMUM_TITLE_TOKEN and word not in TITLE_STOPWORDS
    )


def law_references(value: object) -> frozenset[str]:
    """Riigikogu proceeding numbers named in a title."""
    return frozenset(
        f"{match.group(1)} {match.group(2).upper()}"
        for match in LAW_REFERENCE.finditer(str(value or ""))
    )


#: Ministry abbreviations the two sources do not share, as reviewed pairs.
#:
#: Measured rather than guessed. 163 of the 192 files the reconciliation left
#: `UNMATCHED` have a register row on their own date that this alone kept them
#: from seeing: the archive's filename writes *Majandus- ja
#: Kommunikatsiooniministeerium* and the register's KELLELE writes *MKM*, and
#: `fold` cannot converge an abbreviation with the words it abbreviates.
#:
#: A table rather than a similarity: an abbreviation either is one of these or
#: it is not, and adding a pair is a reviewed act with a name on it. Nothing
#: here is derived from the data at run time, and nothing resolves to an
#: `Organisation` — this is a comparison key, not reference data
#: (app/legacy_import/opinion_recipients.py keeps that boundary).
#:
#: Deliberately absent: *Keskkonnaministeerium* → *Kliimaministeerium*. They
#: look alike and are not the same ministry, and only a reviewed alias may
#: bridge one to the other (docs/adr/0019).
ADDRESSEE_ALIASES: dict[str, str] = {
    "mkm": "majandus ja kommunikatsiooniministeerium",
    "htm": "haridus ja teadusministeerium",
    "rm": "rahandusministeerium",
    "sm": "sotsiaalministeerium",
    "jm": "justiitsministeerium",
    "km": "kliimaministeerium",
    "kkm": "keskkonnaministeerium",
    "rahandusmin": "rahandusministeerium",
    "sotsiaalmin": "sotsiaalministeerium",
    # A missing `e` in the archive's own spelling, not a different ministry.
    "keskkonnaministerium": "keskkonnaministeerium",
    "valitsus": "vabariigi valitsus",
    "regionaalministeerium": "regionaal ja p llumajandusministeerium",
    "rem": "regionaal ja p llumajandusministeerium",
}

#: How a source string separates several addressees, applied to the **raw**
#: text rather than to a folded key.
#:
#: The order matters and cost a test to find: `fold` rewrites every
#: non-alphanumeric character to a space, so a comma is already gone by the
#: time a folded string exists. Splitting the folded form finds nothing to
#: split on and silently returns the whole string as one body — which is the
#: behaviour this was written to replace.
ADDRESSEE_SEPARATOR = re.compile(r"[,;/]|ning", re.IGNORECASE)


def addressee_bodies(value: object) -> frozenset[str]:
    """Every body a recipient string names, as comparison keys.

    Two changes from comparing ``fold(value)`` as one unit, and both are
    measured on the real corpus rather than anticipated.

    **A recipient string may name several bodies.** The archive writes
    ``Siseministeerium, HTM`` where the register writes one of them, so the
    whole-string comparison never fired. Splitting produces a *set*, and the
    comparison becomes "do these two strings name a body in common" — which is
    the question that was always being asked.

    **An abbreviation is the same body as the words it abbreviates**, but only
    where a reviewed pair says so.

    The unsplit string is kept in the set beside the parts, so a register row
    that spells the whole thing out the same way still matches exactly as it
    did. This function can only ever make two strings *more* comparable, never
    less: it is additive over the old key.
    """
    raw = str(value or "")
    folded = fold(raw)
    if not folded:
        return frozenset()
    bodies = {ADDRESSEE_ALIASES.get(folded, folded)}
    for part in ADDRESSEE_SEPARATOR.split(raw):
        piece = fold(part)
        if piece:
            bodies.add(ADDRESSEE_ALIASES.get(piece, piece))
    return frozenset(bodies)


def detect_type(head: bytes) -> str:
    for signature, media_type in SIGNATURES:
        if head.startswith(signature):
            return media_type
    return "application/octet-stream"


@dataclass(frozen=True)
class ArchiveOccurrence:
    """One file at one path inside one archive snapshot."""

    relative_path: str
    original_filename: str
    filename_encoding: str
    sha256: str
    size_bytes: int
    detected_type: str
    filename_date: datetime.date | None
    filename_recipient: str
    filename_title: str

    @property
    def source_year(self) -> int | None:
        return self.filename_date.year if self.filename_date else None


def _decode_entry_name(info: zipfile.ZipInfo) -> tuple[str, str]:
    """Recover a ZIP entry's real name and say how it was recovered.

    ``zipfile`` decodes an entry without general-purpose bit 11 as cp437.
    Re-encoding gets the original bytes back; if they are valid UTF-8 the flag
    was simply missing, which is the case for 91 entries in the real archive.
    If they are not, cp437 was the right reading and is kept.
    """
    if info.flag_bits & 0x800:
        return info.filename, "utf-8-flag"
    raw = info.filename.encode("cp437", errors="replace")
    try:
        return raw.decode("utf-8"), "cp437-bytes-were-utf8"
    except UnicodeDecodeError:
        return info.filename, "cp437"


def _parse_archive_name(name: str) -> tuple[datetime.date | None, str, str]:
    match = ARCHIVE_NAME.match(name)
    if match is None:
        return None, "", ""
    try:
        parsed = datetime.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        # A filename can say 2021-02-30. That is a naming defect, not a date.
        return None, match.group(4).strip(), match.group(5).strip()
    return parsed, match.group(4).strip(), match.group(5).strip()


def _safe_relative_path(name: str) -> str:
    """Refuse anything that could escape a directory it is written into.

    Nothing in this stage extracts to an archive-chosen path, but the guard
    stays where the name is first read: a later caller that does extract must
    not have to remember to re-check (Stage-2H brief 2).

    The rules are deliberately host-independent, because a name that is
    harmless on the machine that reads it can be an absolute path on the
    machine that later writes it: a drive prefix is refused on Linux too, and a
    backslash is refused outright here rather than quietly meaning one thing on
    Windows and another on POSIX. ZIP entries reach this function through
    `_normalize_zip_entry_path`, which translates the separator first.

    A "." segment is refused rather than collapsed. Collapsing it would let
    ``Opinions/./x.pdf`` and ``Opinions/x.pdf`` become one identity while
    remaining two entries, and `archive_relative_path` is part of an
    occurrence's identity.
    """
    if not name or name.startswith("/") or "\\" in name:
        raise OpinionSourceError(f"Refusing an unsafe archive entry name: {name!r}")
    if WINDOWS_DRIVE.match(name):
        raise OpinionSourceError(f"Refusing an unsafe archive entry name: {name!r}")
    if any(part in ("..", ".", "") for part in name.split("/")):
        raise OpinionSourceError(f"Refusing an unsafe archive entry name: {name!r}")
    return posixpath.normpath(name)


def _normalize_zip_entry_path(name: str) -> str | None:
    r"""The one place a ZIP member name becomes an application path.

    The approved archive stores every member as ``Opinions\<name>.pdf``.
    ``zipfile`` replaces the OS separator with ``/`` while parsing a name, so
    that container reads as ``Opinions/<name>.pdf`` on Windows and keeps the
    backslash on Linux — same bytes, same SHA, two different paths, and the
    guard above refused the Linux reading. Translating here makes the canonical
    path the POSIX one on every host.

    The translation happens *before* validation, never after: ``Opinions\..\x``
    has to become ``Opinions/../x`` in time for the traversal check to see it.

    Returns None for a directory record — a name that ends in a separator.
    That question is `zipfile.ZipInfo.is_dir`'s, and it answers it with
    ``os.path.altsep``, so it too reads one stored name as a directory on
    Windows and as a file with an empty last segment on Linux. Answering it
    here, after the translation, is what makes one container hold one set of
    files.
    """
    posix = name.replace("\\", "/")
    if posix.endswith("/"):
        return None
    return _safe_relative_path(posix)


def _canonical_zip_entry(info: zipfile.ZipInfo, seen: dict[str, str]) -> tuple[str, str] | None:
    r"""One entry's canonical path and how its name was decoded, or None.

    None means a directory record, which every caller here skips.

    `seen` maps an already-taken canonical path to the entry name that took it.
    Two members that canonicalise to the same path — ``Opinions/x.pdf`` beside
    ``Opinions\x.pdf``, most plausibly — are refused rather than resolved. Either
    resolution is wrong: as two occurrences they duplicate one identity, and as
    one occurrence a byte sequence silently disappears from the catalogue.
    """
    decoded, encoding = _decode_entry_name(info)
    canonical = _normalize_zip_entry_path(decoded)
    if canonical is None:
        return None
    previous = seen.get(canonical)
    if previous is not None:
        raise OpinionSourceError(
            f"Refusing an archive with two entries at one path: "
            f"{previous!r} and {decoded!r} both mean {canonical!r}."
        )
    seen[canonical] = decoded
    return canonical, encoding


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_opinion_archive(path: Path) -> tuple[str, list[ArchiveOccurrence]]:
    """Inventory the archive. Returns its own SHA-256 and every occurrence.

    Accepts a ZIP or a directory. Neither is modified and nothing is extracted:
    the reconciliation needs hashes and names, and the bytes are read again
    later, once, only for the occurrences that become evidence.
    """
    if not path.exists():
        raise OpinionSourceError(f"{path} does not exist.")
    if path.is_dir():
        return _read_directory(path)
    return file_sha256(path), list(_read_zip(path))


def _read_zip(path: Path) -> Iterator[ArchiveOccurrence]:
    seen: dict[str, str] = {}
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist():
            entry = _canonical_zip_entry(info, seen)
            if entry is None:
                continue
            relative, encoding = entry
            data = archive.read(info)
            base = posixpath.basename(relative)
            parsed_date, recipient, title = _parse_archive_name(base)
            yield ArchiveOccurrence(
                relative_path=relative,
                original_filename=base,
                filename_encoding=encoding,
                sha256=hashlib.sha256(data).hexdigest(),
                size_bytes=len(data),
                detected_type=detect_type(data[:8]),
                filename_date=parsed_date,
                filename_recipient=recipient,
                filename_title=title,
            )


def _read_directory(root: Path) -> tuple[str, list[ArchiveOccurrence]]:
    """A folder input has no archive hash of its own, so one is derived.

    The manifest hash is over (path, sha256) pairs in sorted order, which is the
    same identity a ZIP's hash provides: change any file and the plan will
    refuse to apply (Stage-2H brief 48).
    """
    occurrences: list[ArchiveOccurrence] = []
    for entry in sorted(p for p in root.rglob("*") if p.is_file()):
        relative = entry.relative_to(root).as_posix()
        _safe_relative_path(relative)
        digest = file_sha256(entry)
        with entry.open("rb") as handle:
            head = handle.read(8)
        parsed_date, recipient, title = _parse_archive_name(entry.name)
        occurrences.append(
            ArchiveOccurrence(
                relative_path=relative,
                original_filename=entry.name,
                filename_encoding="filesystem",
                sha256=digest,
                size_bytes=entry.stat().st_size,
                detected_type=detect_type(head),
                filename_date=parsed_date,
                filename_recipient=recipient,
                filename_title=title,
            )
        )
    manifest = "\n".join(f"{o.relative_path}\t{o.sha256}" for o in occurrences)
    return hashlib.sha256(manifest.encode("utf-8")).hexdigest(), occurrences


# ---------------------------------------------------------------------------
# KodaDash
# ---------------------------------------------------------------------------


@dataclass
class KodaDashRow:
    """One producer row, bound to the archive by the producer's own hash."""

    external_id: str
    file_sha256: str
    title: str = ""
    document_date: datetime.date | None = None
    recipient_raw: str = ""
    recipient_normalized: str = ""
    recipient_filter_group: str = ""
    recipient_type: str = ""
    recipient_secondary: str = ""
    recipient_review_required: bool = False
    related_news_url: str = ""
    related_news_id: str = ""
    policy_thread_id: str = ""
    public_import_eligible: bool | None = None
    excluded_from_public: bool = False
    exclusion_reason: str = ""
    payload: dict[str, Any] = field(default_factory=dict)


#: Sheets the reader looks for, and what each contributes. Named rather than
#: positional because the producer's workbook grew from 4 sheets to 34 between
#: revisions and a positional read would have silently changed meaning.
BINDING_SHEET = "02_source_binding_audit"
RECIPIENT_SHEET = "19_recipient_normalization_audit"
IMPORT_SHEET = "opinions_app_import"
EXCLUDED_SHEET = "excluded_rows"

#: Column aliases, because the producer renamed fields between revisions. The
#: first present name wins; a missing column is absent data, not an error.
ALIASES: dict[str, tuple[str, ...]] = {
    "external_id": ("content_id",),
    "file_sha256": ("file_sha256", "sha256"),
    "title": ("title", "title_from_baseline", "title_from_filename"),
    "document_date": ("document_date",),
    "recipient_raw": ("recipient_raw",),
    "recipient_normalized": ("recipient_normalized_after", "recipient", "recipient_normalized"),
    "recipient_filter_group": ("recipient_filter_group_after", "recipient_filter_group"),
    "recipient_type": ("recipient_type_after", "recipient_type"),
    "recipient_secondary": ("recipient_secondary",),
    "recipient_review_required": ("recipient_normalization_review_required",),
    "related_news_url": ("related_koda_news_url",),
    "related_news_id": ("related_koda_news_content_id",),
    "policy_thread_id": ("canonical_policy_thread_id", "policy_thread_id"),
    "public_import_eligible": ("final_app_import_eligible",),
    "exclusion_reason": ("final_import_exclusion_reason", "manual_review_reason"),
}


def _first(row: dict[str, Any], names: tuple[str, ...]) -> Any:
    for name in names:
        value = row.get(name)
        if value not in (None, "", "None"):
            return value
    return None


def _as_bool(value: Any) -> bool | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("true", "1", "yes", "jah", "y"):
        return True
    if text in ("false", "0", "no", "ei", "n"):
        return False
    return None


def _as_date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "").strip()[:10]
    try:
        return datetime.date.fromisoformat(text)
    except ValueError:
        return None


def _sheet_rows(book: Any, name: str) -> list[dict[str, Any]]:
    if name not in book.sheetnames:
        return []
    worksheet = book[name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for values in rows:
        record = dict(zip(header, values, strict=False))
        if any(value not in (None, "") for value in record.values()):
            out.append(record)
    return out


def read_kodadash_artifact(path: Path) -> tuple[str, list[KodaDashRow]]:
    """Read the producer workbook and return its hash and its rows.

    The binding sheet is preferred because it is the only one that records the
    archive SHA-256. Without it there is no deterministic KodaDash-to-archive
    edge at all, and the reconciliation says so rather than falling back to
    filenames (Stage-2H brief 13).
    """
    import openpyxl

    if not path.exists():
        raise OpinionSourceError(f"{path} does not exist.")
    digest = file_sha256(path)
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)

    binding = _sheet_rows(book, BINDING_SHEET)
    if not binding:
        raise OpinionSourceError(
            f"{path.name} carries no {BINDING_SHEET!r} sheet, so no row can be bound to the "
            "archive by hash. Matching KodaDash rows to the archive by filename is not "
            "supported: on the real corpus it produced three collisions and five wrong "
            "assignments where the recorded hashes were exact."
        )

    enrichment: dict[str, dict[str, Any]] = {}
    for sheet in (RECIPIENT_SHEET, IMPORT_SHEET, EXCLUDED_SHEET):
        for record in _sheet_rows(book, sheet):
            key = str(_first(record, ALIASES["external_id"]) or "")
            if key:
                enrichment.setdefault(key, {}).update(
                    {k: v for k, v in record.items() if v not in (None, "", "None")}
                )

    excluded_ids = {
        str(_first(record, ALIASES["external_id"]) or "")
        for record in _sheet_rows(book, EXCLUDED_SHEET)
    }

    rows: list[KodaDashRow] = []
    for record in binding:
        external_id = str(_first(record, ALIASES["external_id"]) or "")
        sha = str(_first(record, ALIASES["file_sha256"]) or "").strip().lower()
        if not external_id or not re.fullmatch(r"[0-9a-f]{64}", sha or ""):
            continue
        merged = dict(record)
        merged.update(enrichment.get(external_id, {}))
        rows.append(
            KodaDashRow(
                external_id=external_id,
                file_sha256=sha,
                title=str(_first(merged, ALIASES["title"]) or ""),
                document_date=_as_date(_first(merged, ALIASES["document_date"])),
                recipient_raw=str(_first(merged, ALIASES["recipient_raw"]) or ""),
                recipient_normalized=str(_first(merged, ALIASES["recipient_normalized"]) or ""),
                recipient_filter_group=str(_first(merged, ALIASES["recipient_filter_group"]) or ""),
                recipient_type=str(_first(merged, ALIASES["recipient_type"]) or ""),
                recipient_secondary=str(_first(merged, ALIASES["recipient_secondary"]) or ""),
                recipient_review_required=bool(
                    _as_bool(_first(merged, ALIASES["recipient_review_required"]))
                ),
                related_news_url=str(_first(merged, ALIASES["related_news_url"]) or ""),
                related_news_id=str(_first(merged, ALIASES["related_news_id"]) or ""),
                policy_thread_id=str(_first(merged, ALIASES["policy_thread_id"]) or ""),
                public_import_eligible=_as_bool(_first(merged, ALIASES["public_import_eligible"])),
                excluded_from_public=external_id in excluded_ids,
                exclusion_reason=str(_first(merged, ALIASES["exclusion_reason"]) or ""),
                payload={
                    key: str(value)
                    for key, value in merged.items()
                    if key and value not in (None, "", "None")
                },
            )
        )
    return digest, rows


class ArchiveReader:
    """Reads one occurrence's bytes, from a ZIP or a directory.

    Opened once for a whole run rather than per file: 767 separate opens of a
    105 MB ZIP is the kind of thing that looks fine on a laptop and takes
    minutes on a server.

    Lives here rather than beside one of its callers because both the canonical
    apply and the archive materialisation need to read the same bytes the same
    way, and two readers would eventually decode a ZIP entry name differently.

    It is keyed by `_canonical_zip_entry`, the same function the inventory uses,
    so the path the catalogue stored is the path that opens the file. The value
    is the `ZipInfo` rather than its name: two members can carry one stored name
    once separators are canonicalised, and a name lookup would then return
    whichever `zipfile` reached first.
    """

    def __init__(self, path: Path) -> None:
        self._path = path
        self._zip: zipfile.ZipFile | None = None
        self._entries: dict[str, zipfile.ZipInfo] = {}
        if path.is_file():
            self._zip = zipfile.ZipFile(path)
            seen: dict[str, str] = {}
            for info in self._zip.infolist():
                entry = _canonical_zip_entry(info, seen)
                if entry is not None:
                    self._entries[entry[0]] = info

    def read(self, relative_path: str) -> bytes | None:
        if self._zip is not None:
            info = self._entries.get(relative_path)
            if info is None:
                return None
            return self._zip.read(info)
        candidate = self._path / relative_path
        if not candidate.is_file():
            return None
        return candidate.read_bytes()
