"""Reading the workbook. No database, no interpretation, no writes.

This module is deliberately usable on a machine with no PostgreSQL, because the
first thing anyone needs to do with a new snapshot is look at it — and requiring
a database to do that would push people back towards opening the real file in
Excel and eyeballing it.

Two design rules earn their keep here:

**openpyxl, not pandas.** A hyperlink is the only surviving pointer to the
OneNote page behind a matter, and the difference between a date cell and a
string that looks like ``43831`` is evidence about which era wrote the row.
pandas discards both on the way to a DataFrame. Here they are the payload.

**Fail closed on structure.** A sheet whose headers do not match its contract is
not parsed with best guesses; it produces findings and its rows are marked for
review. The one thing worse than an unimported year is an imported year whose
columns are one to the left.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.legacy_import.contracts import EraContract, contract_for_year, load_contracts

#: Bumped whenever the reader's output changes shape or meaning. Stored on every
#: batch and every source reference.
PARSER_VERSION = "1.0"

SOURCE_SYSTEM = "excel-register"

_YEAR_SHEET = re.compile(r"^\d{4}$")

#: Reference sheets carry vocabulary, not matters. They are inventoried and used
#: for validation, never parsed as rows.
REFERENCE_SHEETS: frozenset[str] = frozenset({"Hetkeseisu info"})


def serialize_cell_value(value: Any) -> str:
    """A stable text form of a cell value.

    Determinism matters more than prettiness: this string is what gets stored as
    provenance and what an idempotency check compares against, so the same cell
    must produce the same text on every machine and every run. Strings are
    **not** stripped — trailing whitespace is part of what the source said.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # repr is the shortest round-tripping form and is stable across
        # platforms; str() collapses some values that are not equal.
        return repr(value)
    if isinstance(value, str):
        return value
    return str(value)  # pragma: no cover - openpyxl yields no other types


@dataclass(frozen=True)
class SourceCell:
    """One cell, with the metadata that turns a value into evidence."""

    letter: str
    header: str
    raw: str
    value: Any
    data_type: str
    number_format: str
    hyperlink: str

    @property
    def is_blank(self) -> bool:
        return not self.raw.strip()


@dataclass(frozen=True)
class SourceRow:
    sheet: str
    year: int
    row_number: int
    cells: dict[str, SourceCell]

    @property
    def is_blank(self) -> bool:
        return all(cell.is_blank for cell in self.cells.values())

    def cell(self, letter: str) -> SourceCell | None:
        return self.cells.get(letter)

    def text(self, letter: str) -> str:
        cell = self.cells.get(letter)
        return cell.raw if cell is not None else ""

    def raw_mapping(self) -> dict[str, str]:
        """The whole row as ``column letter -> raw text``.

        Keyed by letter rather than header so that a sheet with a blank or
        duplicated header still round-trips, and so the stored provenance does
        not silently change shape when a header is corrected in a later
        snapshot.
        """
        return {letter: cell.raw for letter, cell in sorted(self.cells.items())}

    def hyperlinks(self) -> dict[str, str]:
        return {
            letter: cell.hyperlink for letter, cell in sorted(self.cells.items()) if cell.hyperlink
        }


@dataclass(frozen=True)
class HeaderFinding:
    """A structural disagreement between a sheet and its contract."""

    column: str
    expected: str
    found: str
    kind: str  # "mismatch" | "uncontracted"

    def describe(self) -> str:
        if self.kind == "uncontracted":
            return f"Veerg {self.column}: leping ei kirjelda seda veergu (pealkiri {self.found!r})."
        return (
            f"Veerg {self.column}: leping ootab pealkirja {self.expected!r}, "
            f"leht ütleb {self.found!r}."
        )


@dataclass
class SheetInventory:
    name: str
    is_year_sheet: bool
    year: int | None
    max_row: int
    max_column: int
    contract_version: str = ""
    header_row: int | None = None
    headers: dict[str, str] = field(default_factory=dict)
    header_findings: list[HeaderFinding] = field(default_factory=list)
    data_row_count: int = 0
    blank_row_count: int = 0
    hyperlink_count: int = 0
    uncontracted_columns_with_data: dict[str, int] = field(default_factory=dict)

    @property
    def has_contract(self) -> bool:
        return bool(self.contract_version)

    @property
    def structure_ok(self) -> bool:
        return self.has_contract and not self.header_findings


@dataclass
class WorkbookInventory:
    path: str
    file_name: str
    sha256: str
    byte_size: int
    sheet_names: list[str]
    sheets: list[SheetInventory]
    parser_version: str = PARSER_VERSION

    @property
    def year_sheets(self) -> list[SheetInventory]:
        return [sheet for sheet in self.sheets if sheet.is_year_sheet]

    @property
    def sheets_without_contract(self) -> list[str]:
        return [sheet.name for sheet in self.year_sheets if not sheet.has_contract]


def file_sha256(path: Path) -> tuple[str, int]:
    """Byte-level identity of the snapshot. Streamed, so size is not a factor."""
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
            size += len(chunk)
    return digest.hexdigest(), size


def _hyperlink_of(cell: Cell) -> str:
    link = cell.hyperlink
    if link is None:
        return ""
    target = getattr(link, "target", "") or ""
    location = getattr(link, "location", "") or ""
    if target and location:
        return f"{target}#{location}"
    return target or location


def _headers_of(sheet: Worksheet, header_row: int) -> dict[str, str]:
    headers: dict[str, str] = {}
    for index in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=index).value
        text = "" if value is None else str(value).strip()
        headers[get_column_letter(index)] = text
    return headers


def check_headers(headers: dict[str, str], contract: EraContract) -> list[HeaderFinding]:
    """Compare a sheet's headers with what its contract promises.

    A column the contract does not describe is only a finding if the sheet
    labels it. Trailing unlabelled columns are Excel's own padding — the 2020
    sheet claims thirty-four of them — and reporting each one as a defect would
    bury the findings that matter. Whether such a column nevertheless *holds
    data* is a separate, louder question, answered while the rows are read.
    """
    findings: list[HeaderFinding] = []
    expected = {column.letter: column.header for column in contract.columns}

    for letter, header in expected.items():
        found = headers.get(letter, "")
        if found != header:
            findings.append(
                HeaderFinding(column=letter, expected=header, found=found, kind="mismatch")
            )

    for letter, found in headers.items():
        if letter not in expected and found:
            findings.append(
                HeaderFinding(column=letter, expected="", found=found, kind="uncontracted")
            )
    return findings


def _read_row(
    sheet: Worksheet,
    row_number: int,
    year: int,
    headers: dict[str, str],
    columns: int,
) -> SourceRow:
    cells: dict[str, SourceCell] = {}
    for index in range(1, columns + 1):
        letter = get_column_letter(index)
        cell = sheet.cell(row=row_number, column=index)
        cells[letter] = SourceCell(
            letter=letter,
            header=headers.get(letter, ""),
            raw=serialize_cell_value(cell.value),
            value=cell.value,
            data_type=cell.data_type,
            number_format=cell.number_format or "",
            hyperlink=_hyperlink_of(cell),
        )
    return SourceRow(sheet=sheet.title, year=year, row_number=row_number, cells=cells)


class RegisterWorkbook:
    """An opened snapshot. Read-only, and it never touches the original file."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")
        self.sha256, self.byte_size = file_sha256(self.path)
        # data_only surfaces the values a person saw in the cells rather than
        # the formulas behind them; hyperlinks survive either way.
        self._workbook = load_workbook(self.path, data_only=True, rich_text=False)

    def close(self) -> None:
        self._workbook.close()

    def __enter__(self) -> RegisterWorkbook:
        return self

    def __exit__(self, *exc: object) -> None:
        self.close()

    @property
    def sheet_names(self) -> list[str]:
        return list(self._workbook.sheetnames)

    def year_of(self, sheet_name: str) -> int | None:
        return int(sheet_name) if _YEAR_SHEET.match(sheet_name) else None

    def inventory(self) -> WorkbookInventory:
        """Describe every sheet without interpreting a single matter."""
        contracts = load_contracts()
        sheets: list[SheetInventory] = []

        for name in self.sheet_names:
            worksheet = self._workbook[name]
            year = self.year_of(name)
            inventory = SheetInventory(
                name=name,
                is_year_sheet=year is not None,
                year=year,
                max_row=worksheet.max_row or 0,
                max_column=worksheet.max_column or 0,
            )
            contract = contracts.get(year) if year is not None else None
            if contract is not None:
                inventory.contract_version = contract.contract_version
                inventory.header_row = contract.header_row
                inventory.headers = _headers_of(worksheet, contract.header_row)
                inventory.header_findings = check_headers(inventory.headers, contract)
                self._count_rows(worksheet, contract, inventory)
            sheets.append(inventory)

        return WorkbookInventory(
            path=str(self.path),
            file_name=self.path.name,
            sha256=self.sha256,
            byte_size=self.byte_size,
            sheet_names=self.sheet_names,
            sheets=sheets,
        )

    def _count_rows(
        self, worksheet: Worksheet, contract: EraContract, inventory: SheetInventory
    ) -> None:
        contracted = {column.letter for column in contract.columns}
        for row in self.rows(contract):
            if row.is_blank:
                inventory.blank_row_count += 1
                continue
            inventory.data_row_count += 1
            inventory.hyperlink_count += len(row.hyperlinks())
            for letter, cell in row.cells.items():
                if letter not in contracted and not cell.is_blank:
                    inventory.uncontracted_columns_with_data[letter] = (
                        inventory.uncontracted_columns_with_data.get(letter, 0) + 1
                    )
        _ = worksheet

    def rows(self, contract: EraContract) -> list[SourceRow]:
        """Every row below the header, blanks included and counted."""
        worksheet = self._workbook[contract.sheet]
        headers = _headers_of(worksheet, contract.header_row)
        # Read at least as far as the contract describes, so a column the
        # contract expects is reported as empty rather than silently absent.
        columns = max(worksheet.max_column or 0, contract.last_contracted_index)
        return [
            _read_row(worksheet, number, contract.year, headers, columns)
            for number in range(contract.header_row + 1, (worksheet.max_row or 0) + 1)
        ]

    def reference_vocabulary(self, sheet_name: str = "Hetkeseisu info") -> list[tuple[str, str]]:
        """The controlled ``Hetkeseis`` list, for validation only.

        Read so the importer can report disagreement between the workbook's own
        vocabulary sheet and the seeded ``StageVocabulary``. It never writes:
        overwriting reviewed help text from a sheet that may itself be stale is
        how a vocabulary quietly drifts (Stage-2A brief 17).
        """
        if sheet_name not in self.sheet_names:
            return []
        worksheet = self._workbook[sheet_name]
        entries: list[tuple[str, str]] = []
        for number in range(2, (worksheet.max_row or 0) + 1):
            label = serialize_cell_value(worksheet.cell(row=number, column=1).value).strip()
            if not label:
                continue
            explanation = serialize_cell_value(worksheet.cell(row=number, column=2).value).strip()
            entries.append((label, explanation))
        return entries


def contract_or_none(year: int | None) -> EraContract | None:
    return contract_for_year(year) if year is not None else None
