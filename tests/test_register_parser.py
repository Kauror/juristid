"""Reading the workbook: structure, cells, dates, references and hyperlinks.

Every test here runs against a synthetic workbook generated in a temp directory.
No database is touched, which is also the point — the offline inspector has to
work on a machine that has none.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest

from app.legacy_import.contracts import contract_for_year
from app.legacy_import.dates import (
    RULE_BLANK,
    RULE_NATIVE,
    RULE_SERIAL_STRING,
    RULE_UNPARSED,
    parse_date,
    response_interval_days,
)
from app.legacy_import.enums import Anomaly, RowOutcome
from app.legacy_import.extraction import extract_row, extract_sheet, read_count
from app.legacy_import.inspection import build_summary, classify_offline, inspect_workbook
from app.legacy_import.parser import RegisterWorkbook, serialize_cell_value
from app.legacy_import.references import parse_matter_reference
from tests.synthetic_register import (
    ONENOTE_LINK,
    Row,
    Sheet,
    duplicate_reference_corpus,
    era_corpus,
    pre_numbered_corpus,
    write_workbook,
)


@pytest.fixture
def corpus(tmp_path: Path) -> Path:
    return write_workbook(tmp_path / "synthetic.xlsx", era_corpus())


# -- references ------------------------------------------------------------


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("2026_123", (2026, 123)),
        ("  2011_1  ", (2011, 1)),
        ("2026_12a", None),
        ("2026-12", None),
        ("2026_012", None),  # a different token than 2026_12
        ("2026", None),
        ("", None),
        (None, None),
        ("2026_0", None),
    ],
)
def test_reference_parsing_is_strict(value: str | None, expected: tuple[int, int] | None) -> None:
    parsed = parse_matter_reference(value)
    assert (None if parsed is None else (parsed.year, parsed.number)) == expected


# -- dates -----------------------------------------------------------------


def test_a_real_date_cell_reads_as_itself() -> None:
    parsed = parse_date(dt.datetime(2026, 1, 15))
    assert parsed.value == dt.date(2026, 1, 15)
    assert parsed.rule == RULE_NATIVE


def test_an_excel_serial_typed_as_text_is_still_a_date() -> None:
    """The most common historical anomaly in the pre-2020 sheets."""
    # 40544 is 2011-01-01 on Excel's 1900 system, so 40543 is New Year's Eve.
    parsed = parse_date("40543")
    assert parsed.value == dt.date(2010, 12, 31)
    assert parsed.rule == RULE_SERIAL_STRING
    assert parsed.raw == "40543", "the raw value survives the interpretation"


def test_a_blank_cell_is_blank_and_never_the_epoch() -> None:
    parsed = parse_date(None)
    assert parsed.value is None
    assert parsed.rule == RULE_BLANK
    assert not parsed.failed, "an absent date is not a failed date"


def test_an_unreadable_value_fails_loudly_rather_than_guessing() -> None:
    parsed = parse_date("mitte kuupäev")
    assert parsed.value is None
    assert parsed.rule == RULE_UNPARSED
    assert parsed.failed


def test_an_impossible_calendar_date_is_not_repaired() -> None:
    parsed = parse_date("31.02.2019")
    assert parsed.value is None
    assert parsed.failed


def test_a_bare_number_outside_the_plausible_range_is_not_a_date() -> None:
    assert parse_date("12").failed or parse_date("12").value is None
    assert parse_date(999999).value is None


def test_negative_response_intervals_are_preserved() -> None:
    """Deadlines that had already started running are the department's reality."""
    assert response_interval_days(dt.date(2026, 1, 10), dt.date(2026, 1, 1)) == -9


# -- counts ----------------------------------------------------------------


def test_blank_and_zero_counts_are_different_facts() -> None:
    blank = read_count("")
    zero = read_count("0")
    assert blank.value is None and blank.is_blank
    assert zero.value == 0 and not zero.is_blank


def test_an_unreadable_count_is_not_silently_zero() -> None:
    count = read_count("umbes 20")
    assert count.value is None
    assert not count.readable


# -- serialization ---------------------------------------------------------


def test_cell_serialization_is_deterministic_and_does_not_strip() -> None:
    assert serialize_cell_value(None) == ""
    assert serialize_cell_value("  padded  ") == "  padded  "
    assert serialize_cell_value(dt.datetime(2026, 1, 2, 3, 4)) == "2026-01-02 03:04:00"
    assert serialize_cell_value(dt.date(2026, 1, 2)) == "2026-01-02"
    assert serialize_cell_value(7) == "7"
    assert serialize_cell_value(True) == "true"


# -- workbook structure ----------------------------------------------------


def test_inventory_finds_every_year_sheet_and_its_contract(corpus: Path) -> None:
    with RegisterWorkbook(corpus) as workbook:
        inventory = workbook.inventory()
    assert inventory.sha256 and len(inventory.sha256) == 64
    assert not inventory.sheets_without_contract
    years = sorted(sheet.year for sheet in inventory.year_sheets if sheet.year)
    assert years == [2011, 2017, 2018, 2020, 2021, 2022, 2023, 2024, 2025, 2026]


def test_headers_match_every_contract(corpus: Path) -> None:
    with RegisterWorkbook(corpus) as workbook:
        inventory = workbook.inventory()
    for sheet in inventory.year_sheets:
        assert not sheet.header_findings, f"{sheet.name}: {sheet.header_findings}"


def test_a_shifted_header_is_a_finding_not_a_guess(tmp_path: Path) -> None:
    """The contract fails closed. Importing a year one column to the left is
    the failure mode that produces confident, wrong history."""
    path = write_workbook(tmp_path / "shifted.xlsx", [Sheet(2026, [Row(reference="2026_1")])])

    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook["2026"].cell(row=1, column=7).value = "KELLELT"  # 2026 must say KELLELE
    workbook.save(path)
    workbook.close()

    with RegisterWorkbook(path) as opened:
        inventory = opened.inventory()
    findings = next(s for s in inventory.year_sheets if s.name == "2026").header_findings
    assert findings
    assert "KELLELE" in findings[0].describe()


def test_the_2021_title_row_does_not_become_a_matter(tmp_path: Path) -> None:
    path = write_workbook(
        tmp_path / "y2021.xlsx",
        [Sheet(2021, [Row(reference="2021_1", title="Sünteetiline teema", owner="Kadri")])],
    )
    contract = contract_for_year(2021)
    assert contract is not None
    with RegisterWorkbook(path) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    assert [row.display_reference for row in rows] == ["2021_1"]


def test_a_hyperlink_on_the_title_cell_is_preserved_exactly(corpus: Path) -> None:
    contract = contract_for_year(2026)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    linked = next(row for row in rows if row.display_reference == "2026_1")
    assert linked.onenote_url == ONENOTE_LINK


# -- extraction ------------------------------------------------------------


def test_the_2011_era_reads_its_counterparty_as_the_sender(corpus: Path) -> None:
    contract = contract_for_year(2011)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2011_1")
    assert row.counterparty_direction == "source"
    assert row.counterparty_raw == "Näidisministeerium"


def test_the_2020_era_reads_its_counterparty_as_the_addressee(corpus: Path) -> None:
    contract = contract_for_year(2020)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2020_1")
    assert row.counterparty_direction == "addressee"


def test_the_raw_row_is_kept_whole(corpus: Path) -> None:
    contract = contract_for_year(2026)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2026_1")
    assert row.raw_row["A"] == "2026_1"
    assert row.raw_row["C"] == "seadus"
    assert set(row.raw_row) >= set("ABCDEFGHIJKL")


def test_more_answers_than_asks_is_recorded_not_corrected(corpus: Path) -> None:
    contract = contract_for_year(2021)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2021_1")
    assert row.feedback_responded.value == 9
    assert row.feedback_requested.value == 3
    assert Anomaly.FEEDBACK_RESPONDED_EXCEEDS_REQUESTED.value in row.anomalies


def test_the_unlabelled_2022_column_is_kept_and_flagged(corpus: Path) -> None:
    contract = contract_for_year(2022)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2022_1")
    assert row.unknown_values == {"K": "X"}
    assert Anomaly.UNKNOWN_COLUMN_VALUE.value in row.anomalies


def test_a_free_text_status_is_never_mapped_by_similarity(corpus: Path) -> None:
    contract = contract_for_year(2024)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    row = next(row for row in rows if row.display_reference == "2024_1")
    assert row.status_raw == "Riigikogus 2. lugemisel"
    assert Anomaly.UNMAPPED_STATUS.value in row.anomalies


def test_duplicate_references_flag_every_member_of_the_group(tmp_path: Path) -> None:
    path = write_workbook(tmp_path / "dupes.xlsx", duplicate_reference_corpus())
    contract = contract_for_year(2026)
    assert contract is not None
    with RegisterWorkbook(path) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)
    duplicates = [row for row in rows if Anomaly.DUPLICATE_REFERENCE.value in row.anomalies]
    assert len(duplicates) == 2, "the first occurrence is not automatically the right one"

    invalid = next(row for row in rows if row.reference_raw == "mitte viide")
    assert Anomaly.INVALID_REFERENCE.value in invalid.anomalies


def test_a_reference_only_row_is_a_reserved_number_not_a_broken_matter(tmp_path: Path) -> None:
    """The live 2026 sheet is numbered to 300 with 192 rows in use.

    Calling the other 108 defective would bury the rows that need a decision.
    """
    path = write_workbook(tmp_path / "pre.xlsx", pre_numbered_corpus(filled=2, reserved_to=10))
    contract = contract_for_year(2026)
    assert contract is not None
    with RegisterWorkbook(path) as workbook:
        rows = extract_sheet(workbook.rows(contract), contract)

    reserved = [row for row in rows if row.is_reserved_reference]
    assert len(reserved) == 8
    assert all(not row.anomalies for row in reserved)
    assert all(classify_offline(row) == RowOutcome.RESERVED_REFERENCE.value for row in reserved)


# -- whole-workbook accounting --------------------------------------------


def test_every_row_is_accounted_for_exactly_once(corpus: Path) -> None:
    """The claim the whole migration rests on, as arithmetic."""
    inspection = inspect_workbook(corpus)
    summary = build_summary(inspection)
    assert sum(summary["outcomes"].values()) == summary["totals"]["rows_below_header"]
    assert summary["totals"]["rows_below_header"] == summary["totals"]["blank_rows"] + summary[
        "totals"
    ]["matter_rows"] + summary["totals"]["reserved_references"] + summary["outcomes"].get(
        RowOutcome.NON_MATTER_ROW.value, 0
    )


def test_the_summary_reports_the_snapshot_hash(corpus: Path) -> None:
    summary = build_summary(inspect_workbook(corpus))
    assert len(summary["source"]["sha256"]) == 64


def test_the_workbook_vocabulary_is_compared_but_never_adopted(corpus: Path) -> None:
    summary = build_summary(inspect_workbook(corpus))
    vocabulary = summary["vocabulary"]
    assert vocabulary["workbook_label_count"] == 11
    assert vocabulary["known_label_count"] == 11
    assert vocabulary["labels_missing_from_seed"] == []
    # The 2024 free-text value is used but is not in the controlled list, and
    # it stays that way.
    assert "Riigikogus 2. lugemisel" in vocabulary["used_labels_not_in_controlled_vocabulary"]


def test_a_sheet_without_a_contract_is_reported_and_not_parsed(tmp_path: Path) -> None:
    path = write_workbook(tmp_path / "future.xlsx", [Sheet(2026, [Row(reference="2026_1")])])

    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook.create_sheet(title="2099")
    workbook.save(path)
    workbook.close()

    inspection = inspect_workbook(path)
    summary = build_summary(inspection)
    assert summary["source"]["sheets_without_contract"] == ["2099"]
    assert any("2099" in finding for finding in summary["structure_findings"])


def test_extract_row_needs_no_database(corpus: Path) -> None:
    """Guard on the property that makes local inspection possible at all."""
    contract = contract_for_year(2026)
    assert contract is not None
    with RegisterWorkbook(corpus) as workbook:
        source_rows = workbook.rows(contract)
    assert extract_row(source_rows[0], contract).sheet == "2026"
