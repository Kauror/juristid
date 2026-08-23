"""A synthetic opinions archive and a synthetic producer workbook.

Every byte here is invented. No Koda opinion, member name, ministry
correspondence or real filename may appear in a fixture — the real corpus is
read for analysis and never committed (Stage-2H brief 50, 72).

The shapes are faithful even though the content is not, because the defects
worth testing are shape defects: a ZIP entry whose name is UTF-8 without the
UTF-8 flag, a producer row bound by hash rather than by filename, a bundle of a
letter and its annex sharing one day.
"""

from __future__ import annotations

import datetime
import hashlib
import zipfile
from dataclasses import dataclass
from pathlib import Path

#: Enough of a PDF that the signature sniffer recognises it. Nothing parses it.
PDF_HEADER = b"%PDF-1.4\n"


#: ZIP record signatures. Needed because the UTF-8 name flag has to be cleared
#: after `zipfile` has written it (see `write_archive`).
LOCAL_HEADER = b"PK\x03\x04"
DIRECTORY_ENTRY = b"PK\x01\x02"


def pdf_bytes(marker: str) -> bytes:
    """Distinct, deterministic bytes with a real PDF signature."""
    return PDF_HEADER + f"% {marker}\n".encode() + b"1 0 obj\n<<>>\nendobj\n%%EOF\n"


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


@dataclass(frozen=True)
class SyntheticOpinion:
    """One file destined for the synthetic archive."""

    name: str
    data: bytes
    #: When true the entry is written without the UTF-8 general-purpose flag,
    #: reproducing the 91 real entries `zipfile` mis-decodes as cp437.
    unflagged: bool = False

    @property
    def sha256(self) -> str:
        return sha256(self.data)


def opinion(
    *,
    date: str,
    recipient: str,
    title: str,
    marker: str | None = None,
    unflagged: bool = False,
) -> SyntheticOpinion:
    """Build a file that follows the archive's `YYYY-MM-DD - Saaja - Pealkiri.pdf`."""
    return SyntheticOpinion(
        name=f"Opinions/{date} - {recipient} - {title}.pdf",
        data=pdf_bytes(marker or f"{date}|{recipient}|{title}"),
        unflagged=unflagged,
    )


def write_archive(path: Path, opinions: list[SyntheticOpinion]) -> Path:
    """Write a ZIP, honouring each entry's requested flag state.

    ``zipfile`` sets general-purpose bit 11 for any name it cannot encode as
    ASCII and offers no way to opt out, so an entry that must arrive unflagged
    is written normally and then has its flag cleared in the two places the
    format stores it. That is fiddly, and it is the only way to reproduce the
    91 real entries whose UTF-8 names carry no UTF-8 flag — the case that
    silently mis-decodes into mojibake if the reader does not handle it.
    """
    with zipfile.ZipFile(path, "w", zipfile.ZIP_STORED) as archive:
        for item in opinions:
            archive.writestr(zipfile.ZipInfo(item.name, date_time=(2026, 1, 1, 0, 0, 0)), item.data)
    unflagged = {item.name for item in opinions if item.unflagged}
    if unflagged:
        _clear_utf8_flag(path, unflagged)
    return path


def write_raw_archive(path: Path, entries: list[tuple[str, bytes]]) -> Path:
    r"""A ZIP whose member names are stored exactly as given.

    ``ZipInfo.__init__`` rewrites the OS separator to ``/``, so on Windows
    ``writestr("Opinions\x.pdf", ...)`` silently produces a forward-slash
    member and the defect under test disappears before it is written. Assigning
    ``filename`` after construction skips that rewrite, which is the only way to
    put a real backslash — or any other malformed name — into the container from
    either platform.

    The names here are deliberately hostile. Nothing in this helper validates
    them; refusing them is the reader's job and the point of the tests.
    """
    with zipfile.ZipFile(path, "w", zipfile.ZIP_STORED) as archive:
        for name, data in entries:
            info = zipfile.ZipInfo("placeholder", date_time=(2026, 1, 1, 0, 0, 0))
            info.filename = name
            archive.writestr(info, data)
    return path


def _clear_utf8_flag(path: Path, names: set[str]) -> None:
    """Clear bit 11 in each named entry's local header and directory record.

    Offsets are taken from `zipfile`'s own parse rather than by scanning for
    signatures, because a signature scan can match compressed payload bytes.
    """
    with zipfile.ZipFile(path) as archive:
        local_offsets = [
            info.header_offset for info in archive.infolist() if info.filename in names
        ]
        directory_start = archive.start_dir

    raw = bytearray(path.read_bytes())

    # Local file header: signature(4) version(2) flags(2) … — bit 11 of the
    # little-endian flag word is bit 0x08 of that word's second byte.
    for offset in local_offsets:
        assert bytes(raw[offset : offset + 4]) == LOCAL_HEADER
        raw[offset + 7] &= ~0x08

    # Central directory entry: signature(4) made-by(2) needed(2) flags(2) …
    cursor = directory_start
    while bytes(raw[cursor : cursor + 4]) == DIRECTORY_ENTRY:
        name_length = int.from_bytes(raw[cursor + 28 : cursor + 30], "little")
        extra_length = int.from_bytes(raw[cursor + 30 : cursor + 32], "little")
        comment_length = int.from_bytes(raw[cursor + 32 : cursor + 34], "little")
        start = cursor + 46
        stored = raw[start : start + name_length].decode("utf-8", errors="replace")
        if stored in names:
            raw[cursor + 9] &= ~0x08
        cursor = start + name_length + extra_length + comment_length

    path.write_bytes(bytes(raw))


def write_kodadash_workbook(
    path: Path,
    rows: list[dict[str, object]],
    *,
    excluded: list[dict[str, object]] | None = None,
    include_binding_sheet: bool = True,
) -> Path:
    """A producer workbook shaped like the real one.

    The binding sheet is what carries ``file_sha256``. It can be omitted, which
    is the case the reader must refuse rather than fall back to filenames.
    """
    import openpyxl

    book = openpyxl.Workbook()
    binding = book.active
    binding.title = "02_source_binding_audit" if include_binding_sheet else "01_inventory"
    binding.append(["content_id", "source_file", "zip_path", "file_sha256", "document_date"])
    for row in rows:
        binding.append(
            [
                row["content_id"],
                row.get("source_file", ""),
                row.get("zip_path", ""),
                row.get("file_sha256", ""),
                row.get("document_date", ""),
            ]
        )

    recipients = book.create_sheet("19_recipient_normalization_audit")
    recipients.append(
        [
            "content_id",
            "recipient_raw",
            "recipient_normalized_after",
            "recipient_filter_group_after",
            "recipient_type_after",
            "recipient_secondary",
            "recipient_normalization_review_required",
        ]
    )
    for row in rows:
        recipients.append(
            [
                row["content_id"],
                row.get("recipient_raw", ""),
                row.get("recipient_normalized", ""),
                row.get("recipient_filter_group", ""),
                row.get("recipient_type", ""),
                row.get("recipient_secondary", ""),
                "TRUE" if row.get("recipient_review_required") else "FALSE",
            ]
        )

    app_import = book.create_sheet("opinions_app_import")
    app_import.append(
        [
            "content_id",
            "title",
            "public_summary",
            "chamber_position",
            "business_impact",
            "law_tags_confirmed",
            "topic_primary",
            "related_koda_news_url",
            "related_koda_news_content_id",
            "canonical_policy_thread_id",
            "final_app_import_eligible",
        ]
    )
    for row in rows:
        app_import.append(
            [
                row["content_id"],
                row.get("title", ""),
                row.get("public_summary", ""),
                row.get("chamber_position", ""),
                row.get("business_impact", ""),
                row.get("law_tags_confirmed", ""),
                row.get("topic_primary", ""),
                row.get("related_koda_news_url", ""),
                row.get("related_koda_news_content_id", ""),
                row.get("canonical_policy_thread_id", ""),
                "TRUE" if row.get("public_import_eligible", True) else "FALSE",
            ]
        )

    excluded_sheet = book.create_sheet("excluded_rows")
    excluded_sheet.append(
        ["content_id", "title", "final_app_import_eligible", "final_import_exclusion_reason"]
    )
    for row in excluded or []:
        excluded_sheet.append(
            [row["content_id"], row.get("title", ""), "FALSE", row.get("exclusion_reason", "")]
        )

    book.save(path)
    return path


def iso(value: str) -> datetime.date:
    return datetime.date.fromisoformat(value)
